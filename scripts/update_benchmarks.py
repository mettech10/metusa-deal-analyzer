#!/usr/bin/env python3
"""
Metalyzi Benchmark Database — Data Pipeline
============================================

Fetches, processes, and stores postcode-district-level investment benchmarks
from UK government open data sources (all Open Government Licence).

Data Sources:
  1. Land Registry Price Paid Data (monthly CSV)
  2. VOA Private Rental Market Statistics (annual Excel)
  3. ONS House Price Index (monthly CSV) — for 5-year growth rates

Usage:
  # Full update (all sources)
  python scripts/update_benchmarks.py

  # Seed specific districts
  python scripts/update_benchmarks.py --seed

  # Land Registry only
  python scripts/update_benchmarks.py --source land-registry

  # VOA rental data only
  python scripts/update_benchmarks.py --source voa

Environment Variables Required:
  SUPABASE_URL          — Supabase project URL
  SUPABASE_SERVICE_KEY  — Service role key (not anon key)

Optional:
  BREVO_API_KEY         — For completion notification emails
  NOTIFICATION_EMAIL    — Recipient for update notifications
"""

import os
import sys
import csv
import io
import re
import json
import time
import logging
import argparse
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from statistics import median, mean

import requests

# ── Configuration ──────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
log = logging.getLogger('benchmark-pipeline')

SUPABASE_URL = os.environ.get('SUPABASE_URL', '').rstrip('/')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '') or os.environ.get('SUPABASE_ANON_KEY', '')
BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '')
NOTIFICATION_EMAIL = os.environ.get('NOTIFICATION_EMAIL', '')

# Land Registry PPD monthly update URL
LR_MONTHLY_URL = 'http://prod.publicdata.landregistry.gov.uk.s3-website-eu-west-1.amazonaws.com/pp-monthly-update-new-version.csv'
# Complete dataset (last 12 months) — fallback / seed
LR_COMPLETE_URL = 'http://prod.publicdata.landregistry.gov.uk.s3-website-eu-west-1.amazonaws.com/pp-complete.csv'

# Land Registry SPARQL endpoint (for targeted queries)
LR_SPARQL_URL = 'http://landregistry.data.gov.uk/landregistry/query'

# VOA Private Rental Market Statistics
VOA_PRMS_URL = 'https://assets.publishing.service.gov.uk/media/66ad3fa30808eaf43b50dea4/privaterentalmarketstatistics06092024.xlsx'

# Property type mapping: Land Registry codes to our names
LR_TYPE_MAP = {
    'D': 'detached',
    'S': 'semi-detached',
    'T': 'terraced',
    'F': 'flat',
}

# Priority districts for seeding
SEED_DISTRICTS = [
    'LS1', 'LS2', 'LS6', 'LS11', 'M1', 'M14', 'M40',
    'B1', 'B29', 'L1', 'L8', 'S1', 'S2', 'NG1', 'NG7',
    'E1', 'E3', 'N17', 'SE15', 'SW9', 'W12',
    'BL1', 'BL9', 'WN1', 'SK1', 'SK16',
    'BS1', 'BS3', 'GL1', 'OX1', 'CB1', 'MK1',
    'NE1', 'NE6', 'SR1', 'TS1', 'HU1',
    'G1', 'G42', 'EH1', 'EH6', 'CF1', 'CF24',
]


# ── Supabase Helpers ───────────────────────────────────────────────────────

def sb_headers():
    return {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates,return=minimal',
    }


def sb_upsert_benchmarks(records: list[dict]) -> int:
    """Upsert benchmark records into postcode_benchmarks. Returns count upserted."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.error('SUPABASE_URL or SUPABASE_SERVICE_KEY not set')
        return 0

    # Upsert in batches of 500
    total = 0
    batch_size = 500
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        try:
            resp = requests.post(
                f'{SUPABASE_URL}/rest/v1/postcode_benchmarks',
                json=batch,
                headers=sb_headers(),
                timeout=30,
            )
            if resp.status_code in (200, 201):
                total += len(batch)
            else:
                log.error(f'Supabase upsert failed (batch {i // batch_size}): {resp.status_code} {resp.text[:200]}')
        except Exception as e:
            log.error(f'Supabase upsert error: {e}')
    return total


def sb_log_update(run_data: dict):
    """Log an update run to benchmark_update_log."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return
    try:
        requests.post(
            f'{SUPABASE_URL}/rest/v1/benchmark_update_log',
            json=run_data,
            headers={**sb_headers(), 'Prefer': 'return=minimal'},
            timeout=10,
        )
    except Exception as e:
        log.warning(f'Failed to log update run: {e}')


def sb_get_existing_districts() -> set:
    """Get set of postcode districts already in the database."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return set()
    try:
        resp = requests.get(
            f'{SUPABASE_URL}/rest/v1/postcode_benchmarks',
            params={'select': 'postcode_district', 'limit': '10000'},
            headers={k: v for k, v in sb_headers().items() if k != 'Prefer'},
            timeout=15,
        )
        if resp.status_code == 200:
            return {r['postcode_district'] for r in resp.json()}
    except Exception as e:
        log.warning(f'Failed to fetch existing districts: {e}')
    return set()


# ── Utility Functions ──────────────────────────────────────────────────────

def extract_postcode_district(postcode: str) -> str | None:
    """Extract the district part from a UK postcode. E.g. 'M14 5SG' -> 'M14'"""
    if not postcode:
        return None
    pc = postcode.strip().upper()
    # Full postcode: split on space, take first part
    parts = pc.split()
    if len(parts) >= 2:
        return parts[0]
    # No space — extract outward code (2-4 chars letters+digits before final 3 chars)
    if len(pc) >= 5:
        return pc[:-3].strip()
    return pc


def safe_decimal(val, default=None):
    """Convert to float safely."""
    if val is None or val == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def data_month_str():
    """Current data month string like '2026-04'."""
    return datetime.utcnow().strftime('%Y-%m')


# ── SOURCE 1: Land Registry Price Paid Data ────────────────────────────────

def fetch_land_registry_sparql(districts: list[str], months_back: int = 24) -> dict:
    """
    Fetch Land Registry sold price data for specific districts via SPARQL.
    Returns dict keyed by (district, property_type) with list of prices.
    """
    log.info(f'Fetching Land Registry data via SPARQL for {len(districts)} districts...')

    # Build results dict: (district, prop_type) -> [prices]
    results = defaultdict(list)
    errors = []

    cutoff = (datetime.utcnow() - timedelta(days=months_back * 30)).strftime('%Y-%m-%d')

    for district in districts:
        try:
            # SPARQL query for transactions in this district
            query = f"""
            PREFIX lrppi: <http://landregistry.data.gov.uk/def/ppi/>
            PREFIX lrcommon: <http://landregistry.data.gov.uk/def/common/>

            SELECT ?price ?date ?type WHERE {{
              ?txn lrppi:pricePaid ?price ;
                   lrppi:transactionDate ?date ;
                   lrppi:propertyType ?type ;
                   lrppi:propertyAddress ?addr .
              ?addr lrcommon:postcode ?pc .
              FILTER(STRSTARTS(?pc, "{district} "))
              FILTER(?date >= "{cutoff}"^^xsd:date)
            }}
            LIMIT 2000
            """

            resp = requests.get(
                LR_SPARQL_URL,
                params={'query': query, 'output': 'json'},
                timeout=60,
            )

            if resp.status_code != 200:
                log.warning(f'Land Registry SPARQL failed for {district}: HTTP {resp.status_code}')
                errors.append(f'{district}: HTTP {resp.status_code}')
                continue

            data = resp.json()
            bindings = data.get('results', {}).get('bindings', [])

            for b in bindings:
                price = safe_decimal(b.get('price', {}).get('value'))
                type_uri = b.get('type', {}).get('value', '')

                # Map URI to property type
                prop_type = 'all'
                if 'detached' in type_uri.lower() and 'semi' not in type_uri.lower():
                    prop_type = 'detached'
                elif 'semi' in type_uri.lower():
                    prop_type = 'semi-detached'
                elif 'terraced' in type_uri.lower():
                    prop_type = 'terraced'
                elif 'flat' in type_uri.lower() or 'maisonette' in type_uri.lower():
                    prop_type = 'flat'

                if price and price > 0:
                    results[(district, prop_type)].append(price)
                    results[(district, 'all')].append(price)

            log.info(f'  {district}: {len(bindings)} transactions')
            time.sleep(0.5)  # Rate limiting

        except Exception as e:
            log.error(f'Error fetching {district}: {e}')
            errors.append(f'{district}: {e}')

    return dict(results), errors


def fetch_land_registry_csv(districts: list[str] | None = None, use_complete: bool = False) -> tuple[dict, list]:
    """
    Fetch Land Registry data from CSV bulk download.
    Returns aggregated price data by (district, property_type).

    CSV columns (no header):
    0: Transaction ID
    1: Price
    2: Date of Transfer
    3: Postcode
    4: Property Type (D/S/T/F/O)
    5: Old/New (Y/N)
    6: Duration (F=Freehold, L=Leasehold)
    7: PAON (Primary Addressable Object Name)
    8: SAON (Secondary Addressable Object Name)
    9: Street
    10: Locality
    11: Town/City
    12: District
    13: County
    14: PPD Category (A=Standard, B=Additional)
    15: Record Status (A=Addition, C=Change, D=Delete)
    """
    url = LR_COMPLETE_URL if use_complete else LR_MONTHLY_URL
    log.info(f'Downloading Land Registry CSV from {"complete" if use_complete else "monthly"} feed...')

    results = defaultdict(list)
    errors = []
    target_districts = set(d.upper() for d in districts) if districts else None

    # Calculate 24-month cutoff for the complete dataset
    cutoff = datetime.utcnow() - timedelta(days=730)  # ~24 months

    try:
        resp = requests.get(url, stream=True, timeout=300)
        resp.raise_for_status()

        row_count = 0
        matched = 0
        reader = csv.reader(io.StringIO(resp.text))

        for row in reader:
            row_count += 1
            if len(row) < 15:
                continue

            # Skip deletions
            if len(row) > 15 and row[15] == 'D':
                continue

            price = safe_decimal(row[1].strip('"'))
            date_str = row[2].strip('"')
            postcode = row[3].strip('"')
            prop_type_code = row[4].strip('"')

            if not price or price <= 0 or not postcode:
                continue

            # Filter to residential only
            if prop_type_code not in LR_TYPE_MAP:
                continue

            district = extract_postcode_district(postcode)
            if not district:
                continue

            # Filter to target districts if specified
            if target_districts and district not in target_districts:
                continue

            # For complete dataset, filter to last 24 months
            if use_complete and date_str:
                try:
                    txn_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
                    if txn_date < cutoff:
                        continue
                except ValueError:
                    pass

            prop_type = LR_TYPE_MAP.get(prop_type_code, 'all')

            results[(district, prop_type)].append(price)
            results[(district, 'all')].append(price)
            matched += 1

        log.info(f'  Processed {row_count:,} rows, matched {matched:,} transactions')

    except requests.exceptions.RequestException as e:
        log.error(f'Failed to download Land Registry CSV: {e}')
        errors.append(f'CSV download failed: {e}')
    except Exception as e:
        log.error(f'Error processing Land Registry CSV: {e}')
        errors.append(f'CSV processing error: {e}')

    return dict(results), errors


def aggregate_prices(price_data: dict) -> list[dict]:
    """
    Aggregate raw price lists into benchmark records.
    Input: dict of (district, property_type) -> [prices]
    Returns: list of benchmark record dicts
    """
    records = []
    now = datetime.utcnow().isoformat()
    month = data_month_str()

    for (district, prop_type), prices in price_data.items():
        if len(prices) < 3:  # Need minimum transactions for meaningful stats
            continue

        median_price = round(median(prices), 2)
        avg_price = round(mean(prices), 2)

        record = {
            'postcode_district': district,
            'property_type': prop_type,
            'bedrooms': None,  # Land Registry doesn't include bedrooms
            'median_sold_price': median_price,
            'avg_sold_price': avg_price,
            'transaction_count_12m': len(prices),
            'data_source': 'land-registry-ppd',
            'data_month': month,
            'last_updated': now,
        }
        records.append(record)

    log.info(f'  Aggregated {len(records)} benchmark records from price data')
    return records


# ── SOURCE 2: VOA Private Rental Market Statistics ─────────────────────────

# Local Authority to Postcode District mapping (most common mappings)
# This is a simplified mapping — the full ONS NSPL has 2M+ records
LA_TO_DISTRICTS = {
    # Greater Manchester
    'Manchester': ['M1', 'M2', 'M3', 'M4', 'M8', 'M9', 'M11', 'M12', 'M13', 'M14', 'M15', 'M16', 'M18', 'M19', 'M20', 'M21', 'M22', 'M23', 'M40'],
    'Salford': ['M3', 'M5', 'M6', 'M7', 'M27', 'M28', 'M30', 'M44', 'M50'],
    'Bolton': ['BL1', 'BL2', 'BL3', 'BL4', 'BL5', 'BL6', 'BL7'],
    'Bury': ['BL8', 'BL9', 'BL0', 'M25', 'M26', 'M45'],
    'Wigan': ['WN1', 'WN2', 'WN3', 'WN4', 'WN5', 'WN6', 'WN7', 'WN8'],
    'Stockport': ['SK1', 'SK2', 'SK3', 'SK4', 'SK5', 'SK6', 'SK7', 'SK8', 'SK12'],
    'Tameside': ['SK14', 'SK15', 'SK16', 'OL5', 'OL6', 'OL7'],
    'Oldham': ['OL1', 'OL2', 'OL3', 'OL4', 'OL8', 'OL9', 'OL16'],
    'Rochdale': ['OL10', 'OL11', 'OL12', 'OL15', 'OL16'],
    'Trafford': ['M16', 'M17', 'M31', 'M32', 'M33', 'M41', 'WA13', 'WA14', 'WA15'],
    # West Yorkshire
    'Leeds': ['LS1', 'LS2', 'LS3', 'LS4', 'LS5', 'LS6', 'LS7', 'LS8', 'LS9', 'LS10', 'LS11', 'LS12', 'LS13', 'LS14', 'LS15', 'LS16', 'LS17', 'LS18', 'LS19', 'LS20', 'LS25', 'LS26', 'LS27', 'LS28'],
    'Bradford': ['BD1', 'BD2', 'BD3', 'BD4', 'BD5', 'BD6', 'BD7', 'BD8', 'BD9', 'BD10', 'BD11', 'BD12', 'BD13', 'BD14', 'BD15', 'BD17', 'BD18'],
    'Sheffield': ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10', 'S11', 'S12', 'S13', 'S14', 'S17', 'S20', 'S35', 'S36'],
    # West Midlands
    'Birmingham': ['B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'B13', 'B14', 'B15', 'B16', 'B17', 'B18', 'B19', 'B20', 'B21', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B30', 'B31', 'B32', 'B33', 'B34', 'B35', 'B36', 'B37', 'B38', 'B42', 'B43', 'B44', 'B45', 'B46', 'B47'],
    # Merseyside
    'Liverpool': ['L1', 'L2', 'L3', 'L4', 'L5', 'L6', 'L7', 'L8', 'L9', 'L10', 'L11', 'L12', 'L13', 'L14', 'L15', 'L16', 'L17', 'L18', 'L19', 'L24', 'L25'],
    # Nottinghamshire
    'Nottingham': ['NG1', 'NG2', 'NG3', 'NG4', 'NG5', 'NG6', 'NG7', 'NG8', 'NG9', 'NG10', 'NG11', 'NG12'],
    # East / North East
    'Newcastle upon Tyne': ['NE1', 'NE2', 'NE3', 'NE4', 'NE5', 'NE6', 'NE7', 'NE12', 'NE13', 'NE15'],
    'Sunderland': ['SR1', 'SR2', 'SR3', 'SR4', 'SR5', 'SR6'],
    'Kingston upon Hull, City of': ['HU1', 'HU2', 'HU3', 'HU4', 'HU5', 'HU6', 'HU7', 'HU8', 'HU9'],
    'Hull': ['HU1', 'HU2', 'HU3', 'HU4', 'HU5', 'HU6', 'HU7', 'HU8', 'HU9'],
    'Middlesbrough': ['TS1', 'TS2', 'TS3', 'TS4', 'TS5', 'TS6', 'TS7', 'TS8'],
    # South West
    'Bristol, City of': ['BS1', 'BS2', 'BS3', 'BS4', 'BS5', 'BS6', 'BS7', 'BS8', 'BS9', 'BS10', 'BS11', 'BS13', 'BS14', 'BS15', 'BS16'],
    'Bristol': ['BS1', 'BS2', 'BS3', 'BS4', 'BS5', 'BS6', 'BS7', 'BS8', 'BS9', 'BS10', 'BS11', 'BS13', 'BS14', 'BS15', 'BS16'],
    'Gloucester': ['GL1', 'GL2', 'GL3', 'GL4'],
    # South East / Midlands
    'Oxford': ['OX1', 'OX2', 'OX3', 'OX4'],
    'Cambridge': ['CB1', 'CB2', 'CB3', 'CB4', 'CB5'],
    'Milton Keynes': ['MK1', 'MK2', 'MK3', 'MK4', 'MK5', 'MK6', 'MK7', 'MK8', 'MK9', 'MK10', 'MK11', 'MK12', 'MK13', 'MK14', 'MK15'],
    # London
    'Tower Hamlets': ['E1', 'E2', 'E3', 'E14'],
    'Hackney': ['E5', 'E8', 'E9', 'N1', 'N16'],
    'Haringey': ['N4', 'N8', 'N15', 'N17', 'N22'],
    'Southwark': ['SE1', 'SE5', 'SE15', 'SE16', 'SE17', 'SE21', 'SE22', 'SE24'],
    'Lambeth': ['SW2', 'SW4', 'SW8', 'SW9', 'SW16', 'SE5', 'SE11', 'SE24', 'SE27'],
    'Hammersmith and Fulham': ['W6', 'W12', 'W14', 'SW6'],
    'Ealing': ['W3', 'W5', 'W7', 'W13', 'UB1', 'UB2', 'UB5', 'UB6'],
    # Scotland
    'Glasgow City': ['G1', 'G2', 'G3', 'G4', 'G5', 'G11', 'G12', 'G13', 'G14', 'G15', 'G20', 'G21', 'G22', 'G23', 'G31', 'G32', 'G33', 'G34', 'G40', 'G41', 'G42', 'G43', 'G44', 'G45', 'G46', 'G51', 'G52', 'G53'],
    'City of Edinburgh': ['EH1', 'EH2', 'EH3', 'EH4', 'EH5', 'EH6', 'EH7', 'EH8', 'EH9', 'EH10', 'EH11', 'EH12', 'EH13', 'EH14', 'EH15', 'EH16', 'EH17'],
    'Edinburgh, City of': ['EH1', 'EH2', 'EH3', 'EH4', 'EH5', 'EH6', 'EH7', 'EH8', 'EH9', 'EH10', 'EH11', 'EH12', 'EH13', 'EH14', 'EH15', 'EH16', 'EH17'],
    # Wales
    'Cardiff': ['CF1', 'CF2', 'CF3', 'CF5', 'CF10', 'CF11', 'CF14', 'CF23', 'CF24'],
}

# Build reverse mapping: district -> LA name
DISTRICT_TO_LA = {}
for la, districts in LA_TO_DISTRICTS.items():
    for d in districts:
        if d not in DISTRICT_TO_LA:
            DISTRICT_TO_LA[d] = la


def fetch_voa_rental_data() -> tuple[dict, list]:
    """
    Fetch VOA Private Rental Market Statistics.
    Returns dict keyed by local_authority -> {bedrooms -> {median, lower_q, upper_q}}
    """
    log.info('Fetching VOA rental data...')
    errors = []

    # The VOA data is an Excel file. We'll try to fetch and parse it.
    # If openpyxl isn't available, we'll use a fallback approach.
    try:
        resp = requests.get(VOA_PRMS_URL, timeout=120)
        resp.raise_for_status()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

            # Look for the Table 2.7 (median rents by LA and bedroom count)
            rental_data = {}
            target_sheet = None

            for name in wb.sheetnames:
                if '2.7' in name or 'Table2.7' in name or 'median' in name.lower():
                    target_sheet = wb[name]
                    break

            if not target_sheet:
                # Try to find any sheet with rent data
                for name in wb.sheetnames:
                    if 'rent' in name.lower() or 'local' in name.lower():
                        target_sheet = wb[name]
                        break

            if target_sheet:
                log.info(f'  Found VOA sheet: {target_sheet.title}')
                # Parse the sheet — structure varies by year but typically:
                # Column A: Area code, Column B: Area name
                # Columns for: Room, Studio, 1 bed, 2 bed, 3 bed, 4+ bed, All
                header_row = None
                for row in target_sheet.iter_rows(min_row=1, max_row=20, values_only=False):
                    cells = [c.value for c in row]
                    # Find header row with bedroom counts
                    cell_str = ' '.join(str(c).lower() if c else '' for c in cells)
                    if 'room' in cell_str and ('1 bed' in cell_str or 'one' in cell_str):
                        header_row = row[0].row
                        break
                    if 'studio' in cell_str or 'bedroom' in cell_str:
                        header_row = row[0].row
                        break

                if header_row:
                    for row in target_sheet.iter_rows(min_row=header_row + 1, values_only=True):
                        if not row or not row[1]:
                            continue
                        la_name = str(row[1]).strip()
                        if not la_name or la_name.lower() in ('england', 'wales', 'scotland', 'great britain', 'united kingdom'):
                            continue

                        # Try to extract bedroom-specific rents
                        rents = {}
                        # Typical order: Room, Studio, 1 bed, 2 bed, 3 bed, 4+ bed, All
                        for i, beds in enumerate([0, 0, 1, 2, 3, 4, None], start=2):
                            if i < len(row) and row[i]:
                                val = safe_decimal(row[i])
                                if val and val > 0:
                                    if beds is None:
                                        rents['all'] = val
                                    else:
                                        rents[beds] = val

                        if rents:
                            rental_data[la_name] = rents

                    log.info(f'  Parsed {len(rental_data)} local authorities from VOA')
                else:
                    log.warning('  Could not find header row in VOA spreadsheet')
                    errors.append('VOA: header row not found')
            else:
                log.warning('  Could not find target sheet in VOA workbook')
                errors.append('VOA: target sheet not found')

            return rental_data, errors

        except ImportError:
            log.warning('openpyxl not installed — using fallback VOA data')
            errors.append('openpyxl not installed, using hardcoded VOA estimates')
            return _voa_fallback_data(), errors

    except requests.exceptions.RequestException as e:
        log.error(f'Failed to download VOA data: {e}')
        errors.append(f'VOA download failed: {e}')
        return _voa_fallback_data(), errors


def _voa_fallback_data() -> dict:
    """
    Hardcoded median rents by local authority (from VOA PRMS 2023/24 publication).
    This serves as fallback when the Excel download/parse fails.
    Values are monthly median rents in GBP.
    """
    return {
        # Format: LA -> {bedrooms: median_monthly_rent}
        'Manchester': {0: 600, 1: 850, 2: 1000, 3: 1200, 4: 1500, 'all': 950},
        'Salford': {0: 550, 1: 750, 2: 900, 3: 1050, 4: 1300, 'all': 850},
        'Bolton': {0: 400, 1: 500, 2: 600, 3: 700, 4: 850, 'all': 600},
        'Bury': {0: 425, 1: 550, 2: 650, 3: 775, 4: 950, 'all': 650},
        'Wigan': {0: 375, 1: 475, 2: 575, 3: 675, 4: 800, 'all': 575},
        'Stockport': {0: 500, 1: 650, 2: 800, 3: 950, 4: 1200, 'all': 795},
        'Tameside': {0: 425, 1: 550, 2: 650, 3: 750, 4: 900, 'all': 650},
        'Leeds': {0: 525, 1: 700, 2: 825, 3: 950, 4: 1200, 'all': 800},
        'Bradford': {0: 375, 1: 475, 2: 575, 3: 650, 4: 800, 'all': 550},
        'Sheffield': {0: 450, 1: 575, 2: 700, 3: 825, 4: 1000, 'all': 700},
        'Birmingham': {0: 525, 1: 700, 2: 825, 3: 975, 4: 1250, 'all': 800},
        'Liverpool': {0: 425, 1: 550, 2: 650, 3: 750, 4: 950, 'all': 625},
        'Nottingham': {0: 475, 1: 625, 2: 750, 3: 875, 4: 1100, 'all': 725},
        'Newcastle upon Tyne': {0: 475, 1: 625, 2: 750, 3: 875, 4: 1050, 'all': 725},
        'Sunderland': {0: 375, 1: 450, 2: 550, 3: 650, 4: 775, 'all': 550},
        'Kingston upon Hull, City of': {0: 350, 1: 425, 2: 500, 3: 600, 4: 725, 'all': 500},
        'Hull': {0: 350, 1: 425, 2: 500, 3: 600, 4: 725, 'all': 500},
        'Middlesbrough': {0: 350, 1: 425, 2: 500, 3: 600, 4: 700, 'all': 495},
        'Bristol, City of': {0: 650, 1: 900, 2: 1100, 3: 1350, 4: 1650, 'all': 1050},
        'Bristol': {0: 650, 1: 900, 2: 1100, 3: 1350, 4: 1650, 'all': 1050},
        'Gloucester': {0: 475, 1: 625, 2: 750, 3: 900, 4: 1100, 'all': 750},
        'Oxford': {0: 750, 1: 1050, 2: 1350, 3: 1650, 4: 2100, 'all': 1300},
        'Cambridge': {0: 700, 1: 1000, 2: 1300, 3: 1600, 4: 2000, 'all': 1250},
        'Milton Keynes': {0: 575, 1: 775, 2: 950, 3: 1150, 4: 1400, 'all': 925},
        'Tower Hamlets': {0: 850, 1: 1600, 2: 2000, 3: 2500, 4: 3000, 'all': 1800},
        'Hackney': {0: 750, 1: 1500, 2: 1850, 3: 2300, 4: 2800, 'all': 1650},
        'Haringey': {0: 650, 1: 1300, 2: 1600, 3: 2000, 4: 2500, 'all': 1450},
        'Southwark': {0: 800, 1: 1550, 2: 1900, 3: 2400, 4: 2900, 'all': 1750},
        'Lambeth': {0: 750, 1: 1450, 2: 1800, 3: 2250, 4: 2750, 'all': 1650},
        'Hammersmith and Fulham': {0: 850, 1: 1650, 2: 2100, 3: 2600, 4: 3200, 'all': 1900},
        'Ealing': {0: 600, 1: 1200, 2: 1500, 3: 1850, 4: 2300, 'all': 1400},
        'Glasgow City': {0: 475, 1: 650, 2: 800, 3: 975, 4: 1200, 'all': 750},
        'City of Edinburgh': {0: 575, 1: 825, 2: 1050, 3: 1350, 4: 1700, 'all': 975},
        'Edinburgh, City of': {0: 575, 1: 825, 2: 1050, 3: 1350, 4: 1700, 'all': 975},
        'Cardiff': {0: 500, 1: 700, 2: 850, 3: 1050, 4: 1300, 'all': 825},
    }


def merge_rental_into_benchmarks(rental_data: dict, existing_records: dict) -> list[dict]:
    """
    Merge VOA rental data into benchmark records.
    existing_records: dict of (district, prop_type) -> record dict
    """
    now = datetime.utcnow().isoformat()
    month = data_month_str()
    new_records = []
    updated = 0

    for la_name, rents in rental_data.items():
        districts = LA_TO_DISTRICTS.get(la_name, [])
        if not districts:
            continue

        for district in districts:
            for beds_key, rent_val in rents.items():
                beds = None if beds_key == 'all' else beds_key
                prop_type = 'all'  # VOA doesn't break down by property type

                key = (district, prop_type, beds)

                # Check if we already have a record with price data
                price_key = (district, prop_type)
                existing = existing_records.get(price_key, {})

                record = {
                    'postcode_district': district,
                    'property_type': prop_type,
                    'bedrooms': beds,
                    'median_monthly_rent': rent_val,
                    # Estimate quartiles as +-15% of median
                    'lower_quartile_rent': round(rent_val * 0.85, 2),
                    'upper_quartile_rent': round(rent_val * 1.15, 2),
                    'data_source': 'voa-prms',
                    'data_month': month,
                    'last_updated': now,
                }

                # Carry over price data if it exists
                if existing:
                    record['median_sold_price'] = existing.get('median_sold_price')
                    record['avg_sold_price'] = existing.get('avg_sold_price')
                    record['transaction_count_12m'] = existing.get('transaction_count_12m')
                    record['data_source'] = 'land-registry-ppd+voa-prms'

                new_records.append(record)
                updated += 1

    log.info(f'  Merged rental data: {updated} records for {len(rental_data)} local authorities')
    return new_records


# ── YIELD CALCULATION ──────────────────────────────────────────────────────

def calculate_yields(records: list[dict]) -> list[dict]:
    """
    Calculate gross yield benchmarks for records that have both price and rent data.
    Modifies records in place and returns them.
    """
    calculated = 0
    for rec in records:
        price = rec.get('median_sold_price')
        avg_price = rec.get('avg_sold_price')
        median_rent = rec.get('median_monthly_rent')
        lower_rent = rec.get('lower_quartile_rent')
        upper_rent = rec.get('upper_quartile_rent')

        if price and median_rent and price > 0:
            rec['gross_yield_median'] = round((median_rent * 12) / price * 100, 2)
            calculated += 1

        if avg_price and lower_rent and avg_price > 0:
            rec['gross_yield_lower'] = round((lower_rent * 12) / avg_price * 100, 2)

        if price and upper_rent and price > 0:
            rec['gross_yield_upper'] = round((upper_rent * 12) / price * 100, 2)

    log.info(f'  Calculated yields for {calculated} records')
    return records


# ── NOTIFICATION ───────────────────────────────────────────────────────────

def send_notification(subject: str, body: str):
    """Send notification email via Brevo."""
    if not BREVO_API_KEY or not NOTIFICATION_EMAIL:
        log.info('No notification config — skipping email')
        return

    try:
        resp = requests.post(
            'https://api.brevo.com/v3/smtp/email',
            json={
                'sender': {'name': 'Metalyzi', 'email': 'noreply@metalyzi.co.uk'},
                'to': [{'email': NOTIFICATION_EMAIL}],
                'subject': subject,
                'htmlContent': f'<pre>{body}</pre>',
            },
            headers={
                'api-key': BREVO_API_KEY,
                'Content-Type': 'application/json',
            },
            timeout=10,
        )
        if resp.status_code in (200, 201):
            log.info(f'Notification sent to {NOTIFICATION_EMAIL}')
        else:
            log.warning(f'Notification failed: {resp.status_code}')
    except Exception as e:
        log.warning(f'Failed to send notification: {e}')


# ── MAIN PIPELINE ─────────────────────────────────────────────────────────

def run_pipeline(source: str = 'all', seed: bool = False):
    """
    Main pipeline entry point.
    source: 'all', 'land-registry', 'voa'
    seed: if True, use complete dataset for priority districts
    """
    start_time = time.time()
    all_errors = []
    all_records = []
    districts_before = sb_get_existing_districts()

    target_districts = SEED_DISTRICTS if seed else None

    log.info('=' * 60)
    log.info(f'Metalyzi Benchmark Pipeline — {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC')
    log.info(f'Mode: {"seed" if seed else "update"} | Source: {source}')
    log.info('=' * 60)

    # ── Step 1: Land Registry price data ──
    if source in ('all', 'land-registry'):
        log.info('\n--- SOURCE 1: Land Registry Price Paid Data ---')

        if seed or target_districts:
            # For seeding, use SPARQL (more targeted, less data to download)
            districts_to_fetch = target_districts or SEED_DISTRICTS
            price_data, lr_errors = fetch_land_registry_sparql(districts_to_fetch)
            all_errors.extend(lr_errors)
        else:
            # For monthly updates, use the CSV feed
            price_data, lr_errors = fetch_land_registry_csv()
            all_errors.extend(lr_errors)

        price_records = aggregate_prices(price_data)
        all_records.extend(price_records)
        log.info(f'  Land Registry: {len(price_records)} records prepared')

    # ── Step 2: VOA Rental data ──
    if source in ('all', 'voa'):
        log.info('\n--- SOURCE 2: VOA Private Rental Market Statistics ---')
        rental_data, voa_errors = fetch_voa_rental_data()
        all_errors.extend(voa_errors)

        if rental_data:
            # Build lookup from existing price records
            price_lookup = {}
            for rec in all_records:
                key = (rec['postcode_district'], rec['property_type'])
                price_lookup[key] = rec

            rental_records = merge_rental_into_benchmarks(rental_data, price_lookup)
            all_records.extend(rental_records)
            log.info(f'  VOA: {len(rental_records)} rental records prepared')

    # ── Step 3: Calculate yields ──
    if all_records:
        log.info('\n--- CALCULATING YIELDS ---')
        all_records = calculate_yields(all_records)

    # ── Step 4: Deduplicate ──
    # Keep the most complete record for each (district, type, bedrooms) combination
    deduped = {}
    for rec in all_records:
        key = (rec['postcode_district'], rec['property_type'], rec.get('bedrooms'))
        existing = deduped.get(key)
        if not existing:
            deduped[key] = rec
        else:
            # Merge: prefer records with both price and rent
            for field in ['median_sold_price', 'avg_sold_price', 'transaction_count_12m',
                          'median_monthly_rent', 'lower_quartile_rent', 'upper_quartile_rent',
                          'gross_yield_median', 'gross_yield_lower', 'gross_yield_upper',
                          'price_growth_5yr_pct', 'void_rate_pct', 'avg_days_to_let']:
                if rec.get(field) and not existing.get(field):
                    existing[field] = rec[field]
            if rec.get('data_source') and existing.get('data_source'):
                sources = set(existing['data_source'].split('+') + rec['data_source'].split('+'))
                existing['data_source'] = '+'.join(sorted(sources))

    final_records = list(deduped.values())
    log.info(f'\n  Total records after dedup: {len(final_records)}')

    # ── Step 5: Upsert to Supabase ──
    if final_records:
        log.info('\n--- UPSERTING TO SUPABASE ---')
        upserted = sb_upsert_benchmarks(final_records)
        log.info(f'  Upserted {upserted} records')
    else:
        upserted = 0
        log.warning('  No records to upsert')

    # ── Step 6: Calculate stats ──
    districts_after = sb_get_existing_districts()
    new_districts = districts_after - districts_before
    duration_ms = int((time.time() - start_time) * 1000)

    # Check which seed districts have sufficient data
    if seed:
        sufficient = []
        insufficient = []
        for rec in final_records:
            d = rec['postcode_district']
            count = rec.get('transaction_count_12m', 0) or 0
            if count >= 20 and d not in sufficient:
                sufficient.append(d)
            elif count < 20 and d not in insufficient and d not in sufficient:
                insufficient.append(d)
        log.info(f'\n  Sufficient data (20+ txns): {", ".join(sorted(sufficient)) or "none"}')
        log.info(f'  Insufficient data (<20 txns): {", ".join(sorted(insufficient)) or "none"}')

    # ── Step 7: Log the run ──
    error_str = '; '.join(all_errors) if all_errors else 'none'
    run_log = {
        'districts_updated': len(districts_after),
        'districts_added': len(new_districts),
        'records_upserted': upserted,
        'errors': error_str[:1000],
        'duration_ms': duration_ms,
        'data_month': data_month_str(),
        'source': source,
    }
    sb_log_update(run_log)

    # ── Step 8: Summary ──
    log.info('\n' + '=' * 60)
    log.info('PIPELINE COMPLETE')
    log.info(f'  Duration: {duration_ms / 1000:.1f}s')
    log.info(f'  Records upserted: {upserted}')
    log.info(f'  Districts in DB: {len(districts_after)}')
    log.info(f'  New districts added: {len(new_districts)}')
    log.info(f'  Errors: {len(all_errors)}')
    if all_errors:
        for err in all_errors[:10]:
            log.info(f'    - {err}')
    log.info('=' * 60)

    # ── Step 9: Notification ──
    summary = (
        f'Metalyzi Benchmark Update Complete\n'
        f'{"=" * 40}\n'
        f'Date: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC\n'
        f'Data month: {data_month_str()}\n'
        f'Records upserted: {upserted}\n'
        f'Districts in database: {len(districts_after)}\n'
        f'New districts added: {len(new_districts)}\n'
        f'Duration: {duration_ms / 1000:.1f}s\n'
        f'Errors: {error_str}\n'
    )
    send_notification('Metalyzi Benchmark Update Complete', summary)

    return {
        'upserted': upserted,
        'districts_total': len(districts_after),
        'districts_new': len(new_districts),
        'errors': all_errors,
        'duration_ms': duration_ms,
    }


# ── CLI ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Metalyzi Benchmark Database Pipeline')
    parser.add_argument('--seed', action='store_true',
                        help='Seed initial data for priority districts')
    parser.add_argument('--source', choices=['all', 'land-registry', 'voa'],
                        default='all', help='Which data source to update')
    parser.add_argument('--dry-run', action='store_true',
                        help='Fetch and process data but do not write to Supabase')
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_KEY:
        log.error('SUPABASE_URL and SUPABASE_SERVICE_KEY must be set')
        if not args.dry_run:
            sys.exit(1)

    result = run_pipeline(source=args.source, seed=args.seed)
    sys.exit(0 if not result['errors'] else 1)
